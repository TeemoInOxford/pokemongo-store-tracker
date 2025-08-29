import json
import re
from datetime import datetime, timezone, timedelta
import pandas as pd
from opencc import OpenCC
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 售价映射表（非金币）
idr_price_map = {
    5000: 8.99,
    10000: 9.99,
    15000: 12.99,
    25000: 22.22,
    30500: 25.55,
    41000: 29.99,
    50000: 33.33,
    61000: 44.44,
    76000: 55.55,
    101000: 59.99,
    126000: 66.66,
    151000: 77.77,
    200000: 99.99,
    505000: 219.99
    
}

# 金币定价表（只有金币商品用）
coin_price_map = {
    600: 18.99,
    1300: 29.99,
    2700: 54.99,
    5600: 104.99,
    15500: 249.99
}

# 判断金币商品
def is_coin_item(category, name_en):
    return category.startswith("POKECOIN") or "coin" in name_en.lower()

def get_selling_price(category, idr, coin_quantity):
    if category == "POKECOINS":
        return f"{coin_price_map.get(coin_quantity, '未知定价')}"
    else:
        return f"{idr_price_map.get(int(idr), f'{idr:.2f}（未定价）')}"

# 图片保存目录
IMG_DIR = "images"
os.makedirs(IMG_DIR, exist_ok=True)

# 下载图片函数
def download_image(url, filename=None):
    if not filename:
        filename = os.path.basename(url)
    local_path = os.path.join(IMG_DIR, filename)
    if not os.path.exists(local_path):
        try:
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                with open(local_path, "wb") as f:
                    f.write(r.content)
                print(f"✅ Downloaded: {filename}")
            else:
                print(f"❌ Failed: {url}")
        except Exception as e:
            print(f"❌ Error downloading {url}: {e}")
    return local_path

cc = OpenCC('t2s')

def convert_time(ms):
    if not ms:
        return None
    dt = datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc) + timedelta(hours=8)
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def extract_key_parts(localization_key):
    raw_key = localization_key.replace("sku.name.", "")
    match = re.match(r"^([a-zA-Z0-9\-_]+)(?:\.(\w+))?$", raw_key)
    if match:
        return match.group(1), match.group(2)
    return raw_key, None

def resolve_name(map_dict, main_key, sub_key):
    val = map_dict.get(main_key, "")
    if isinstance(val, str):
        return val
    elif isinstance(val, dict):
        return val.get(sub_key, "") if sub_key else list(val.values())[0]
    return ""

with open("data.json", "r", encoding="utf-8") as f:
    data = json.load(f)

lang_data = data.get("zh")
i18n = lang_data["pageProps"].get("_nextI18Next", {}).get("initialI18nStore", {})
name_map_en = i18n.get("en", {}).get("common", {}).get("sku", {}).get("name", {})
name_map_zh = i18n.get("zh-Hant", {}).get("common", {}).get("sku", {}).get("name", {})

categories = lang_data["pageProps"]["catalog"]["webstoreCategoriesList"]
items_info = []

category_map = {
    "TICKETS": "门票礼盒",
    "BUNDLE": "道具礼盒",
    "LIMITED_TIME": "限时礼盒",
    "ITEMS": "道具",
    "POKECOINS": "宝可币",
    "POKECOIN": "宝可币"
}

def suggest_price(idr, premium=0.15):
    base = 0.00048 * idr + 10
    bundle_price = base * (1 + premium)
    return round(bundle_price) - 0.01 if bundle_price > 10 else round(bundle_price, 2)

for category in categories:
    raw_cat = category["category"]
    cat_cn = category_map.get(raw_cat, raw_cat)
    cat_name = f"{raw_cat}\n{cat_cn}"
    for item in category.get("itemsList", []):
        imageUrl = item.get("imageUrl")
        localization_key = item.get("localizationNameKey", "")
        main_key, sub_key = extract_key_parts(localization_key)

        name_en = resolve_name(name_map_en, main_key, sub_key) or main_key
        name_zh_trad = resolve_name(name_map_zh, main_key, sub_key)
        name_zh_simp = cc.convert(name_zh_trad)
        full_name = f"{name_en}\n{name_zh_simp}"
        
        # 🚫 跳过 110 PokéCoins
        if name_en.strip() == "110 PokéCoins":
            continue

        end_time = item.get("webstoreLimitInfo", {}).get("endTimeMs")
        bundle_coin = item.get("bundledCurrencyList", [])
        bundle = item.get("bundledItemList", [])
        price = item.get("priceList")
        idr_price = int(price[0].get("priceE6", 0)) / 1_000_000
        coin_quantity = bundle_coin[0].get("quantity") if raw_cat == "POKECOINS" and bundle_coin else 0
        selling_price = get_selling_price(raw_cat, idr_price, coin_quantity)

        # 下载主图
        if imageUrl:
            image_filename = f"main_{main_key.lower()}.png"
            download_image(imageUrl, image_filename)
            local_main_image_path = f"{IMG_DIR}/{image_filename}"
        else:
            local_main_image_path = ""

        # 打包道具内容
        bundle_info = []
        icon_ids = set()
        for b in bundle:
            raw_id = b.get("itemId")
            b_main_key, b_sub_key = extract_key_parts(raw_id)
            b_name_en = resolve_name(name_map_en, b_main_key, b_sub_key) or raw_id
            b_name_zh_trad = resolve_name(name_map_zh, b_main_key, b_sub_key)
            b_name_zh_simp = cc.convert(b_name_zh_trad)
            bundle_info.append({
                "itemId": f"{b_name_en}\n{b_name_zh_simp}",
                "rawId": raw_id.lower(),
                "quantity": b.get("quantity")
            })
            icon_ids.add(raw_id.lower())

        # 打包货币内容
        for b in bundle_coin:
            currency_id = b.get("currency").lower()
            bundle_info.append({
                "itemId": b.get("currency") + "\n" + category_map.get(b.get("currency"), b.get("currency")),
                "rawId": currency_id,
                "quantity": b.get("quantity")
            })
            icon_ids.add(currency_id)

        # 下载所有小图
        for rid in icon_ids:
            icon_url = f"https://storage.googleapis.com/platform-webstore-rel-assets/pgo/sku_assets/{rid}.png"
            icon_filename = f"icon_{rid}.png"
            download_image(icon_url, icon_filename)

        items_info.append({
            "raw_category": raw_cat,  # 添加原始类别用于排序
            "category": cat_name,
            "name": full_name,
            "endTimeMs": convert_time(end_time),
            "bundledItems": bundle_info,
            "sellingPrice": f"{selling_price} CNY 人民币",
            "price": price,
            "imageUrl": imageUrl,
            "localImage": local_main_image_path
        })

# 创建DataFrame
df = pd.DataFrame(items_info)

# 定义排序顺序
category_order = {
    "BUNDLE": 1,      # 道具礼盒 - 第一组，排在前
    "ITEMS": 2,       # 道具 - 第一组，排在后
    "LIMITED_TIME": 3, # 限时礼盒 - 第二组，排在前
    "TICKETS": 4,     # 门票礼盒 - 第二组，排在后
    "POKECOINS": 5,   # 宝可币 - 不需要DLC
    "POKECOIN": 5     # 宝可币 - 不需要DLC
}

# 排序
df['sort_order'] = df['raw_category'].map(category_order)
df = df.sort_values('sort_order')

# 分配DLC编号
dlc_counter = {
    "group1": 1,  # BUNDLE和ITEMS
    "group2": 1   # LIMITED_TIME和TICKETS
}

def assign_dlc(row):
    if row['raw_category'] in ['POKECOINS', 'POKECOIN']:
        return ""  # 金币不需要DLC
    elif row['raw_category'] in ['BUNDLE', 'ITEMS']:
        dlc = dlc_counter["group1"]
        dlc_counter["group1"] += 1
        return dlc
    elif row['raw_category'] in ['LIMITED_TIME', 'TICKETS']:
        dlc = dlc_counter["group2"]
        dlc_counter["group2"] += 1
        return dlc
    return ""

df['DLC'] = df.apply(assign_dlc, axis=1)

# 删除辅助列
df = df.drop(['raw_category', 'sort_order'], axis=1)

# 重新排列列，DLC在最前面
cols = ['DLC'] + [col for col in df.columns if col != 'DLC']
df = df[cols]

# 导出CSV（不含颜色）
df.to_csv("webstore_items_limited.csv", index=False, encoding="utf-8-sig")

# 导出带颜色的Excel文件
with pd.ExcelWriter("webstore_items_limited.xlsx", engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # 定义颜色
    color_group1 = PatternFill(start_color="B4E7CE", end_color="B4E7CE", fill_type="solid")  # 浅绿色
    color_group2 = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")  # 浅橙色
    
    # 应用颜色到DLC列
    for row in range(2, len(df) + 2):  # Excel行从1开始，第1行是标题
        cell = worksheet.cell(row=row, column=1)  # DLC在第1列
        category_cell = worksheet.cell(row=row, column=2)  # category在第2列
        
        # 根据category判断颜色
        if "道具礼盒" in category_cell.value or "道具" in category_cell.value:
            if "道具礼盒" not in category_cell.value or "道具" in category_cell.value:
                cell.fill = color_group1
        elif "限时礼盒" in category_cell.value or "门票礼盒" in category_cell.value:
            cell.fill = color_group2
    
    # 调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# 导出JSON
df.to_json("webstore_items_limited.json", orient="records", force_ascii=False, indent=2)

print("✅ 导出完成")
print("📊 DLC分配说明：")
print("   - 道具礼盒(BUNDLE)和道具(ITEMS)：使用浅绿色标注")
print("   - 限时礼盒(LIMITED_TIME)和门票礼盒(TICKETS)：使用浅橙色标注")
print("   - 宝可币(POKECOINS)：无DLC编号")
print("📁 已生成文件：")
print("   - webstore_items_limited.csv (无颜色)")
print("   - webstore_items_limited.xlsx (带颜色标注)")
print("   - webstore_items_limited.json")
